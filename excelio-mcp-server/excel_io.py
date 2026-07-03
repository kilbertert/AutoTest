# Copyright (c) Microsoft Corporation.
# Licensed under the MIT license.

"""Excel reader/writer for .xlsx test-case blueprints.

Reading: uses zipfile + xml.etree directly to bypass openpyxl's stylesheet
parser, which crashes on the example 测试用例.xlsx (broken fills node).
Writing: uses openpyxl (we control the file we write, so no broken styles).

All writes are guarded by:
  1. portalocker file lock (cross-process)
  2. a column whitelist — update_cells refuses to write design-time columns
"""

from __future__ import annotations

import os
import shutil
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from xml.etree import ElementTree as ET

import portalocker

NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
# Relationship namespace for workbook.xml.rels
RELS_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"


# ─── column whitelist for update_cells ───────────────────────────────────────
# Blueprint columns (0-indexed):
#   0-13: design-time (用例ID..编写人) — NEVER writable via update_cells
#   14:   执行结果  — writable
#   15:   备注      — writable
#   16:   UI_selector — design-time, NEVER writable
#   17:   截图路径  — writable
WRITABLE_COLS = {14, 15, 17}


class ExcelError(Exception):
    pass


# ─── low-level zip+xml reader ────────────────────────────────────────────────

def _load_shared_strings(zf: zipfile.ZipFile) -> List[str]:
    """Return the shared-strings table as a list of strings, or [] if absent."""
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    # Each <si> is one shared string; its text is the concatenation of <t> children.
    strings: List[str] = []
    for si in root.iter(NS + "si"):
        text = "".join((t.text or "") for t in si.iter(NS + "t"))
        strings.append(text)
    return strings


def _sheet_xml_name_for(zf: zipfile.ZipFile, sheet_name_or_index) -> str:
    """Resolve a user-supplied sheet name or 1-based index to its xl/worksheets/sheetN.xml path."""
    # Coerce numeric strings to int — FastMCP passes unannotated params as str,
    # so an LLM sending sheet=1 arrives as "1" and would otherwise be treated as
    # a (non-matching) sheet name rather than a 1-based index.
    if isinstance(sheet_name_or_index, str) and sheet_name_or_index.strip().isdigit():
        sheet_name_or_index = int(sheet_name_or_index)
    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    sheets: List[Tuple[str, str]] = []  # (name, rId)
    for s in wb_root.iter(NS + "sheet"):
        name = s.attrib.get("name", "")
        rid = s.attrib.get(RELS_NS + "id") or s.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
        sheets.append((name, rid))
    # Resolve by index or name
    target_name: Optional[str] = None
    if isinstance(sheet_name_or_index, int):
        if sheet_name_or_index < 1 or sheet_name_or_index > len(sheets):
            raise ExcelError(f"sheet index {sheet_name_or_index} out of range (1..{len(sheets)})")
        target_name = sheets[sheet_name_or_index - 1][0]
    else:
        for name, _ in sheets:
            if name == sheet_name_or_index:
                target_name = name
                break
        if target_name is None:
            raise ExcelError(f"sheet {sheet_name_or_index!r} not found; available: {[n for n,_ in sheets]}")
    # rId → target file via workbook.xml.rels
    rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rels: Dict[str, str] = {}
    for rel in rels_root:
        rid = rel.attrib.get("Id", "")
        target = rel.attrib.get("Target", "")
        rels[rid] = target
    for name, rid in sheets:
        if name == target_name:
            target = rels.get(rid, "")
            if not target:
                raise ExcelError(f"no relationship target for sheet {target_name!r} (rId={rid})")
            # Target is relative to the workbook's location (xl/). Normalize to
            # an absolute archive path "xl/worksheets/sheetN.xml". openpyxl
            # writes "worksheets/sheet1.xml" (relative); some writers write
            # "/xl/worksheets/..." or "xl/worksheets/...". Handle all.
            target = target.lstrip("/")
            if target.startswith("xl/"):
                return target
            if target.startswith("worksheets/"):
                return "xl/" + target
            # Fallback: assume it's already a full archive path
            return target
    raise ExcelError(f"sheet {target_name!r} not found in rels")


def _parse_sheet(xml_bytes: bytes, shared_strings: List[str], max_rows: Optional[int] = None) -> Tuple[List[List[str]], int, int]:
    """Parse one worksheet XML. Returns (rows, max_row_seen, max_col_seen).

    Each row is a list of cell strings, padded to max_col_seen width.
    """
    root = ET.fromstring(xml_bytes)
    sheet_data = root.find(NS + "sheetData")
    rows_out: List[List[str]] = []
    max_col = 0
    row_count = 0
    if sheet_data is None:
        return rows_out, 0, 0
    for row_el in sheet_data.findall(NS + "row"):
        if max_rows is not None and row_count >= max_rows:
            break
        cells: List[str] = []
        for c in row_el.findall(NS + "c"):
            t = c.attrib.get("t", "")
            v = c.find(NS + "v")
            inline = c.find(NS + "is")
            s = ""
            if v is not None:
                if t == "s":
                    try:
                        s = shared_strings[int(v.text)]
                    except (ValueError, IndexError):
                        s = v.text or ""
                else:
                    s = v.text or ""
            elif inline is not None:
                # inline string: concat <t> children
                s = "".join((t.text or "") for t in inline.iter(NS + "t"))
            cells.append(s)
        if cells:
            max_col = max(max_col, len(cells))
        rows_out.append(cells)
        row_count += 1
    # pad each row to max_col
    for r in rows_out:
        while len(r) < max_col:
            r.append("")
    return rows_out, len(rows_out), max_col


def _col_letters_to_index(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


# ─── public reader API ───────────────────────────────────────────────────────

def list_sheets(path: str) -> List[Dict[str, Any]]:
    """List sheets with name, row count, col count. Does not parse styles."""
    with zipfile.ZipFile(path) as zf:
        shared = _load_shared_strings(zf)
        wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
        out: List[Dict[str, Any]] = []
        for s in wb_root.iter(NS + "sheet"):
            name = s.attrib.get("name", "")
            # sheetId is not the file index; we still report it for reference
            sheet_id = s.attrib.get("sheetId")
            out.append({"name": name, "sheetId": sheet_id, "index": len(out) + 1})
        # To get row/col counts we'd need to open each sheet; do it lazily
        for entry in out:
            try:
                xml_name = _sheet_xml_name_for(zf, entry["index"])
                _, rows, cols = _parse_sheet(zf.read(xml_name), shared, max_rows=None)
                entry["rows"] = rows
                entry["cols"] = cols
            except Exception as e:
                entry["rows"] = 0
                entry["cols"] = 0
                entry["error"] = str(e)
        return out


def read_header(path: str, sheet) -> List[str]:
    """Return the first row of `sheet` (name or 1-based index) as a list of strings."""
    with zipfile.ZipFile(path) as zf:
        shared = _load_shared_strings(zf)
        xml_name = _sheet_xml_name_for(zf, sheet)
        rows, _, _ = _parse_sheet(zf.read(xml_name), shared, max_rows=1)
        if not rows:
            return []
        return rows[0]


def read_sheet(path: str, sheet, row_range: Optional[Dict[str, int]] = None) -> List[Dict[str, Any]]:
    """Read rows from `sheet`. Returns [{row: <1-based>, values: [...]}, ...].

    row_range: {"start": <1-based, inclusive>, "end": <1-based, inclusive>}
              or {"max_rows": <int>} — defaults to all rows.
    The header row (row 1) is included unless `skip_header` semantics are applied by caller.
    """
    with zipfile.ZipFile(path) as zf:
        shared = _load_shared_strings(zf)
        xml_name = _sheet_xml_name_for(zf, sheet)
        all_rows, total_rows, _ = _parse_sheet(zf.read(xml_name), shared, max_rows=None)
    start = 1
    end = total_rows
    if row_range:
        if "max_rows" in row_range:
            start = 1
            end = min(total_rows, row_range["max_rows"])
        else:
            start = max(1, row_range.get("start", 1))
            end = min(total_rows, row_range.get("end", total_rows))
    out: List[Dict[str, Any]] = []
    for i in range(start - 1, end):
        if i >= len(all_rows):
            break
        out.append({"row": i + 1, "values": all_rows[i]})
    return out


def get_module_map(path: str, sheet) -> List[Dict[str, Any]]:
    """Read a 3-column sheet (模块/功能/子功能) as a flat list of dicts.

    Empty parent cells cascade down (Excel-style merged-cell behavior):
    if a row's col 0 is empty, inherit the most recent non-empty col 0 above.
    Same for col 1 → col 2.
    """
    rows = read_sheet(path, sheet)
    if not rows:
        return []
    # Skip the header row (row 1)
    out: List[Dict[str, Any]] = []
    last_module = ""
    last_function = ""
    for r in rows[1:]:
        vals = r["values"]
        module = vals[0] if len(vals) > 0 else ""
        function = vals[1] if len(vals) > 1 else ""
        subfunction = vals[2] if len(vals) > 2 else ""
        if module:
            last_module = module
            last_function = ""  # new module resets function cascade
        else:
            module = last_module
        if function:
            last_function = function
        else:
            function = last_function
        if not (module or function or subfunction):
            continue
        out.append({
            "row": r["row"],
            "module": module,
            "function": function,
            "subfunction": subfunction,
        })
    return out


# ─── public writer API (openpyxl) ────────────────────────────────────────────

def _acquire_lock(path: str):
    """Lock a sibling .lock file. Returns the file handle to pass to release."""
    lock_path = path + ".lock"
    fh = open(lock_path, "a+")
    portalocker.lock(fh, portalocker.LOCK_EX)
    return fh


def _release_lock(fh) -> None:
    try:
        portalocker.unlock(fh)
    finally:
        fh.close()


def create_blueprint(path: str, template_header: List[str], extra_header: Optional[List[str]] = None) -> Dict[str, Any]:
    """Create a new empty .xlsx with the given header row.

    The blueprint has one sheet named "测试用例蓝图" with header = template_header + extra_header.
    Uses openpyxl so we control the styles (no broken stylesheet).
    """
    # Lazy import to keep reader path openpyxl-free
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    p = Path(path).expanduser()
    p.parent.mkdir(parents=True, exist_ok=True)
    if p.exists():
        # Don't clobber an existing blueprint silently — caller should pick a fresh run_id
        raise ExcelError(f"blueprint already exists at {p}; pick a new run_id or delete it first")

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例蓝图"
    header = list(template_header) + list(extra_header or [])
    ws.append(header)
    # Style header row
    bold = Font(bold=True)
    fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Freeze header
    ws.freeze_panes = "A2"

    fh = _acquire_lock(str(p))
    try:
        wb.save(str(p))
    finally:
        _release_lock(fh)

    return {"path": str(p), "columns": len(header), "header": header}


def update_cells(path: str, sheet, updates: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Update cells in `sheet`. Each update: {row: <1-based>, col: <0-based>, value: <str>}.

    Enforces WRITABLE_COLS — only col 14, 15, 17 (执行结果/备注/截图路径) are writable.

    Implementation: pure zip+xml edit of sheet1.xml — never goes through openpyxl.
    This preserves the source workbook's stylesheet (which is often broken in
    xlsx files exported from older Excel versions, crashing openpyxl on load)
    AND preserves Excel-style cascading empty cells (which openpyxl discards
    on save, dropping rows that look empty to it).
    """
    import zipfile
    import xml.etree.ElementTree as ET
    import shutil

    p = Path(path).expanduser()
    if not p.exists():
        raise ExcelError(f"file not found: {p}")

    # Validate whitelist up front
    rejected = []
    accepted = []
    for u in updates:
        col = int(u.get("col", -1))
        if col not in WRITABLE_COLS:
            rejected.append({"row": u.get("row"), "col": col, "reason": f"col {col} is not writable (only {sorted(WRITABLE_COLS)} are)"})
        else:
            accepted.append(u)
    if rejected:
        raise ExcelError(f"update refused: {len(rejected)} cell(s) target non-writable columns: {rejected[:3]}")

    # Resolve sheet xml path. For our 16-col layouts we always use sheet1.
    if isinstance(sheet, int):
        sheet_path = f"xl/worksheets/sheet{sheet}.xml"
    else:
        sheet_path = "xl/worksheets/sheet1.xml"

    fh = _acquire_lock(str(p))
    tmp = p.with_suffix(".xlsx.tmp")
    try:
        with zipfile.ZipFile(p) as zin:
            sheet_bytes = zin.read(sheet_path)
            ss_bytes = zin.read("xl/sharedStrings.xml") if "xl/sharedStrings.xml" in zin.namelist() else None

        ss_root = ET.fromstring(ss_bytes) if ss_bytes else None
        shared_strings = []
        if ss_root is not None:
            for si in ss_root.iter(NS + "si"):
                t = si.find(NS + "t")
                if t is not None:
                    shared_strings.append(t.text or "")
                else:
                    shared_strings.append("".join((tt.text or "") for tt in si.iter(NS + "t")))

        sheet_root = ET.fromstring(sheet_bytes)
        sheet_data = sheet_root.find(NS + "sheetData")
        if sheet_data is None:
            raise ExcelError("no sheetData in target sheet")

        rows_by_idx = {}
        for r in sheet_data.findall(NS + "row"):
            r_attr = r.attrib.get("r", "")
            try:
                rows_by_idx[int(r_attr)] = r
            except ValueError:
                continue

        def col_letter(idx):
            n = idx + 1
            s = ""
            while n > 0:
                n, rem = divmod(n - 1, 26)
                s = chr(65 + rem) + s
            return s

        for u in accepted:
            row_num = int(u["row"])
            col_num = int(u["col"])
            value = str(u.get("value", ""))
            letter = col_letter(col_num)
            ref = f"{letter}{row_num}"

            row = rows_by_idx.get(row_num)
            if row is None:
                row = ET.SubElement(sheet_data, NS + "row")
                row.set("r", str(row_num))
                rows_by_idx[row_num] = row

            cell = None
            for c in row.findall(NS + "c"):
                if c.attrib.get("r", "").startswith(letter):
                    cell = c
                    break
            if cell is None:
                cell = ET.SubElement(row, NS + "c")
                cell.set("r", ref)

            for child in list(cell):
                if child.tag in (NS + "v", NS + "is"):
                    cell.remove(child)
            cell.set("t", "inlineStr")
            is_el = ET.SubElement(cell, NS + "is")
            t_el = ET.SubElement(is_el, NS + "t")
            t_el.text = value

        new_sheet_bytes = ET.tostring(sheet_root, xml_declaration=True, encoding="UTF-8")

        with zipfile.ZipFile(p) as zin:
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.namelist():
                    if item == sheet_path:
                        zout.writestr(item, new_sheet_bytes)
                    else:
                        zout.writestr(item, zin.read(item))
        shutil.move(str(tmp), str(p))
    finally:
        _release_lock(fh)

    return {"written": len(accepted), "rejected": 0}


def append_rows(path: str, sheet, rows: List[List[Any]]) -> Dict[str, Any]:
    """Append rows to `sheet`. Each row is a list of cell values (0-indexed cols).

    Used at design time to add test cases to the blueprint.
    """
    from openpyxl import load_workbook

    p = Path(path).expanduser()
    if not p.exists():
        raise ExcelError(f"file not found: {p}")

    fh = _acquire_lock(str(p))
    try:
        wb = load_workbook(str(p))
        ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet - 1]
        start_row = ws.max_row + 1
        for row in rows:
            ws.append([("" if v is None else str(v)) for v in row])
        wb.save(str(p))
    finally:
        _release_lock(fh)

    return {"appended": len(rows), "start_row": start_row}
